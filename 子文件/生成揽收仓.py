from openpyxl import Workbook, load_workbook

import pandas as pd
import ast  # 用于将字符串安全地转换为列表

# 读取 .xlsx 文件中的特定 sheet

sheet_name = "线路简单"  # 将这里的 "Sheet1" 替换为你要读取的工作表名称

df = pd.read_excel("输入.xlsx", sheet_name=sheet_name)

# 将 path 字段的字符串转换回列表

df["path"] = df["path"].apply(ast.literal_eval)

# 将 DataFrame 转换回字典列表

routes = df.to_dict('records')


# 定义最终生成表的表头
headers = ["揽收仓", "揽收时长（小时）", "开始时间", "开始日期", "完成时间", "完成日期", "元/公斤", "元/票"]

# 创建新的工作簿
wb = Workbook()
ws = wb.active

# 写入表头
ws.append(headers)

try:
    # 加载输入文件
    input_wb = load_workbook('输入.xlsx')
    # 获取揽收仓 sheet
    input_ws = input_wb['揽收仓']

    # 获取输入表的数据（跳过表头）
    input_data = []
    for row in input_ws.iter_rows(min_row=2, values_only=True):
        input_data.append(row)

    # 按照 routes 索引长度重复填入数据
    for _ in range(len(routes)):
        for row in input_data:
            ws.append(row)

    # 保存新文件
    wb.save('ut.xlsx')
    print("数据已处理并保存到 'ut.xlsx' 文件中。")
except FileNotFoundError:
    print("未找到 '输入.xlsx' 文件，请检查文件是否存在。")
except KeyError:
    print("'输入.xlsx' 文件中未找到 '揽收仓' sheet，请检查文件内容。")